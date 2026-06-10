#!/usr/bin/env python3
"""
境外资本利得税计算工具
计算中国大陆税务居民境外证券投资所得的个人所得税

数据源：
- 富途(Futu): Excel (.xlsx) - 港股交易
- 老虎(Tiger): PDF - 美股/港股交易

税务规则：
- 应纳税额 = (境外股转让收入 – 股票买入成本 – 合理税费) × 20%
- 分批买入采用加权平均法计算成本
- 股息/分红收入单独按 20% 计税
- 境外已缴税款可抵扣（分国家/地区计算限额）
"""

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pdfplumber

# ============================================================
# 配置
# ============================================================

# 税率
TAX_RATE = 0.20

# 期初持仓的实际买入成本（用户可在此覆盖自动检测的期初市值）
# 格式: {"代码": 总成本(HKD)} 或 {"代码": (数量, 总成本)}
# 如果不设置，将使用2024年12月31日的市值作为成本基础
INITIAL_COSTS = {
    # "6099": 28800.00,  # 1800股 × 实际买入均价
    # "2459": 808.00,    # 4000股 × 实际买入均价
    # "2460": 2348.00,   # 200股 × 实际买入均价
}


# ============================================================
# 数据模型
# ============================================================


@dataclass
class Trade:
    """股票/证券交易记录"""
    code: str           # 股票代码
    name: str           # 股票名称
    direction: str      # 买入/卖出
    quantity: float     # 数量
    price: float        # 价格
    amount: float       # 成交金额
    fee: float          # 总费用
    net_amount: float   # 变动金额（含费用后净额）
    time: datetime      # 成交时间
    currency: str       # 币种
    market: str         # 交易所/市场
    category: str       # 品类：证券/期权/基金


@dataclass
class Dividend:
    """分红/股息记录"""
    code: str           # 股票代码
    name: str           # 股票名称
    amount: float       # 分红金额
    currency: str       # 币种
    date: datetime      # 日期
    tax_withheld: float = 0.0  # 已扣税


@dataclass
class Position:
    """持仓与加权平均成本"""
    code: str
    name: str
    quantity: float = 0.0
    total_cost: float = 0.0  # 含费用的总成本

    @property
    def avg_cost(self) -> float:
        if self.quantity == 0:
            return 0.0
        return self.total_cost / self.quantity


@dataclass
class RealizedGain:
    """已实现损益"""
    code: str
    name: str
    gain: float         # 已实现损益（原币种）
    currency: str
    market: str         # HK/US
    category: str       # 股票/期权/期权行权


@dataclass
class TigerSummary:
    """老虎账户汇总"""
    # 期权交易
    option_trade_hk: float = 0.0      # HKD
    option_trade_us: float = 0.0      # USD
    # 期权行权/过期
    option_exercise_hk: float = 0.0   # HKD
    option_exercise_us: float = 0.0   # USD
    # 股票
    stock_hk: float = 0.0             # HKD
    stock_us: float = 0.0             # USD
    # 基金（货币基金，可选是否计入）
    fund_hk: float = 0.0              # HKD
    fund_us: float = 0.0              # USD
    # 分红
    dividends: list = field(default_factory=list)
    # 逐笔明细
    stock_details: list = field(default_factory=list)
    option_details: list = field(default_factory=list)
    exercise_details: list = field(default_factory=list)


# ============================================================
# 富途 Excel 解析器
# ============================================================


def parse_futu_trades(filepath: Path) -> tuple[list[Trade], list[Dividend], dict]:
    """解析富途年度账单 Excel
    Returns: (trades, dividends, initial_positions)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    trades = parse_futu_trading_sheet(wb)
    dividends = parse_futu_dividend_sheet(wb)
    initial_positions = parse_futu_initial_positions(wb)
    wb.close()
    return trades, dividends, initial_positions


def parse_futu_initial_positions(wb) -> dict[str, Position]:
    """解析 证券-持仓总览 sheet 中的期初持仓"""
    ws = wb["证券-持仓总览"]
    positions = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        period_type = str(row[0]).strip()
        category = str(row[2] or "").strip()
        code = str(row[5] or "").strip()
        quantity = float(row[8] or 0)
        price = float(row[9] or 0)

        # 只处理期初的证券持仓
        if period_type == "期初" and category == "证券":
            # 使用期初市值作为成本基础（用户可在 INITIAL_COSTS 中覆盖）
            positions[code] = Position(
                code=code,
                name=code,
                quantity=quantity,
                total_cost=quantity * price,  # 期初市值作为默认成本
            )

    # 允许用户通过 INITIAL_COSTS 覆盖实际买入成本
    for code, cost_info in INITIAL_COSTS.items():
        if code in positions:
            positions[code].total_cost = cost_info
        elif isinstance(cost_info, tuple):
            # (quantity, total_cost) 格式
            positions[code] = Position(
                code=code, name=code,
                quantity=cost_info[0], total_cost=cost_info[1]
            )

    return positions


def parse_futu_trading_sheet(wb) -> list[Trade]:
    """解析 证券-交易流水 sheet"""
    ws = wb["证券-交易流水"]
    trades = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        # 列: 成交时间, 账户名称, 账户号码, 品类, 代码名称, 交易所/市场,
        #     方向, 交收日期, 币种, 数量/面值, 价格, 成交金额, 总费用, 变动金额
        time_val = str(row[0]).strip()
        category = str(row[3] or "").strip()
        code_name = str(row[4] or "").strip()
        market = str(row[5] or "").strip()
        direction = str(row[6] or "").strip()
        currency = str(row[8] or "").strip()
        quantity = abs(float(row[9] or 0))
        price = abs(float(row[10] or 0))
        amount = float(row[11] or 0)
        fee = float(row[12] or 0)
        net_amount = float(row[13] or 0)

        # 解析时间
        if isinstance(row[0], datetime):
            trade_time = row[0]
        else:
            try:
                trade_time = datetime.strptime(time_val, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                trade_time = datetime.strptime(time_val[:10], "%Y-%m-%d")

        # 代码名称列实际只包含代码（如 "6099"），不含名称
        code = code_name
        name = code_name  # 名称暂用代码代替

        # 判断买卖方向
        if "买入" in direction:
            dir_simplified = "买入"
        elif "卖出" in direction:
            dir_simplified = "卖出"
        elif direction in ("申购", "赎回"):
            dir_simplified = direction
        else:
            dir_simplified = direction

        trades.append(Trade(
            code=code,
            name=name,
            direction=dir_simplified,
            quantity=quantity,
            price=price,
            amount=amount,
            fee=fee,
            net_amount=net_amount,
            time=trade_time,
            currency=currency,
            market=market,
            category=category,
        ))

    return trades


def parse_futu_dividend_sheet(wb) -> list[Dividend]:
    """解析 证券-资金进出 sheet 中的分红记录

    同时扣除分红相关的费用（Scrip Charge, Handling Charge）
    """
    ws = wb["证券-资金进出"]
    dividends = []
    dividend_fees = {}  # code -> total fees

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        # 列: 日期, 账户名称, 账户号码, 类型, 方向, 币种, 变动金额, 备注
        date_val = str(row[0] or "").strip()
        type_val = str(row[3] or "").strip()
        direction = str(row[4] or "").strip()
        currency = str(row[5] or "").strip()
        amount = float(row[6] or 0)
        remark = str(row[7] or "").strip()

        if type_val != "公司行动":
            continue

        if direction == "In":
            # 分红收入
            code, name = parse_dividend_remark(remark)

            if isinstance(row[0], datetime):
                div_date = row[0]
            else:
                try:
                    div_date = datetime.strptime(date_val, "%Y%m%d")
                except ValueError:
                    try:
                        div_date = datetime.strptime(date_val, "%Y-%m-%d")
                    except ValueError:
                        div_date = datetime(2025, 1, 1)

            # H股分红10%预扣税
            tax_withheld = 0.0
            if "-10%" in remark or "10%" in remark:
                tax_withheld = amount / 9.0

            dividends.append(Dividend(
                code=code,
                name=name,
                amount=amount,
                currency=currency,
                date=div_date,
                tax_withheld=tax_withheld,
            ))

        elif direction == "Out":
            # 分红相关费用（Scrip Charge / Handling Charge）
            if "Scrip Charge" in remark or "Handling Charge" in remark:
                match = re.search(r"<SEHK\s+(\d+)", remark)
                if match:
                    code = match.group(1)
                    dividend_fees[code] = dividend_fees.get(code, 0) + abs(amount)

    # 从分红中扣除关联费用（净分红 = 分红 - 相关费用）
    for d in dividends:
        if d.code in dividend_fees:
            d.amount -= dividend_fees[d.code]
            dividend_fees.pop(d.code)  # 避免重复扣除

    return dividends


def parse_dividend_remark(remark: str) -> tuple[str, str]:
    """从分红备注中提取股票代码和名称
    格式: "24 F/D-HKD0.05/SH <SEHK 732 TRULY INT'L> 10000 shares"
    """
    # 匹配 <SEHK CODE NAME> 模式
    match = re.search(r"<SEHK\s+(\d+)\s+(.+?)>", remark)
    if match:
        return match.group(1), match.group(2)
    # Fallback: 尝试匹配其他格式
    match = re.search(r"(\d{1,5})\s+([A-Z][\w\s'&.-]+)", remark)
    if match:
        return match.group(1), match.group(2).strip()
    return "", remark


# ============================================================
# 富途资本利得计算（加权平均成本法）
# ============================================================


def calculate_futu_gains(
    trades: list[Trade],
    initial_positions: dict[str, Position] = None,
) -> tuple[list[RealizedGain], dict]:
    """
    使用加权平均成本法计算富途股票的已实现损益

    Args:
        trades: 交易记录
        initial_positions: 期初持仓（代码 -> Position）

    Returns:
        realized_gains: 每笔卖出的已实现损益
        positions: 期末持仓
    """
    if initial_positions is None:
        initial_positions = {}

    # 按品类分类
    stock_trades = [t for t in trades if t.category == "证券"]
    option_trades = [t for t in trades if t.category == "期权"]

    # 计算股票损益（加权平均成本法，含期初持仓）— 同时返回期末持仓
    stock_gains, positions = calculate_weighted_avg_gains(stock_trades, "股票", initial_positions)

    # 计算期权损益（按合约净额法）
    option_gains = calculate_option_gains(option_trades)

    all_gains = stock_gains + option_gains

    return all_gains, positions


def calculate_option_gains(trades: list[Trade]) -> list[RealizedGain]:
    """
    计算期权已实现损益

    期权交易逻辑：
    - 卖出开仓：收到权利金（amount > 0），创建空头仓位
    - 买入平仓：支付权利金（amount < 0），关闭空头仓位
    - 期权到期未行权：卖出的权利金全部为利润

    对每个合约，已实现损益 = Σ net_amount（所有已平仓交易的净现金流之和）
    """
    # 按合约代码分组
    codes = set(t.code for t in trades)
    realized_gains = []

    for code in sorted(codes):
        code_trades = sorted(
            [t for t in trades if t.code == code],
            key=lambda x: x.time
        )

        # 计算净持仓量
        net_qty = 0
        total_net_amount = 0.0
        for t in code_trades:
            if t.direction == "卖出":  # 卖出开仓
                net_qty -= t.quantity
                total_net_amount += t.net_amount  # 正值（收入）
            elif t.direction == "买入":  # 买入平仓
                net_qty += t.quantity
                total_net_amount += t.net_amount  # 负值（支出）

        # 如果仓位已平（或已过期），全部为已实现损益
        if abs(net_qty) < 0.001:
            # 完全平仓
            realized_gains.append(RealizedGain(
                code=code,
                name=code,
                gain=total_net_amount,
                currency=code_trades[0].currency,
                market="HK",
                category="期权",
            ))
        else:
            # 仍有未平仓位（可能是到期等待行权或尚未到期）
            # 检查期权是否已到期（到期日在合约代码中）
            expiry = parse_option_expiry(code)
            if expiry and expiry < datetime(2026, 1, 1):
                # 已过期，全部实现
                realized_gains.append(RealizedGain(
                    code=code,
                    name=code,
                    gain=total_net_amount,
                    currency=code_trades[0].currency,
                    market="HK",
                    category="期权",
                ))
            else:
                # 未到期，暂不计入已实现（但记录为待确认）
                realized_gains.append(RealizedGain(
                    code=code,
                    name=f"{code} (未平仓)",
                    gain=total_net_amount,
                    currency=code_trades[0].currency,
                    market="HK",
                    category="期权",
                ))

    return realized_gains


def parse_option_expiry(code: str) -> Optional[datetime]:
    """从期权代码中解析到期日
    格式如: TCH250320C480000 → 2025-03-20
    """
    match = re.search(r"(\d{6})[CP]", code)
    if match:
        date_str = match.group(1)
        try:
            return datetime.strptime("20" + date_str[:2] + date_str[2:4] + date_str[4:6], "%Y%m%d")
        except ValueError:
            pass
    return None


def calculate_weighted_avg_gains(
    trades: list[Trade],
    category: str,
    initial_positions: dict[str, Position] = None,
) -> tuple[list[RealizedGain], dict[str, Position]]:
    """加权平均成本法计算已实现损益（含期初持仓）

    Returns:
        realized_gains: 已实现损益列表
        positions: 期末持仓字典
    """
    if initial_positions is None:
        initial_positions = {}

    # 按股票代码分组
    codes = set(t.code for t in trades)
    # 也包含有期初持仓但本期有交易的代码
    codes.update(k for k in initial_positions if any(t.code == k for t in trades))
    realized_gains = []
    positions = {}

    for code in sorted(codes):
        code_trades = sorted(
            [t for t in trades if t.code == code],
            key=lambda x: x.time
        )

        # 初始化持仓（含期初持仓）
        if code in initial_positions:
            init = initial_positions[code]
            pos = Position(
                code=code, name=init.name,
                quantity=init.quantity, total_cost=init.total_cost
            )
        else:
            pos = Position(code=code, name=code_trades[0].name if code_trades else code)

        for t in code_trades:
            if t.direction == "买入":
                # 买入：成本 = 成交金额绝对值 + 费用
                buy_cost = abs(t.amount) + abs(t.fee)
                pos.quantity += t.quantity
                pos.total_cost += buy_cost
            elif t.direction == "卖出":
                if pos.quantity > 0:
                    # 卖出：收益 = 卖出金额 - 加权平均成本 - 卖出费用
                    sell_qty = min(t.quantity, pos.quantity)
                    avg_cost_per_unit = pos.total_cost / pos.quantity
                    cost_of_sold = avg_cost_per_unit * sell_qty
                    sell_revenue = abs(t.amount)
                    sell_fee = abs(t.fee)

                    gain = sell_revenue - cost_of_sold - sell_fee

                    realized_gains.append(RealizedGain(
                        code=code,
                        name=t.name,
                        gain=gain,
                        currency=t.currency,
                        market="HK",
                        category=category,
                    ))

                    # 更新持仓
                    pos.total_cost -= cost_of_sold
                    pos.quantity -= sell_qty

        positions[code] = pos

    return realized_gains, positions


# ============================================================
# 老虎 PDF 解析器
# ============================================================


def parse_tiger_pdf(filepath: Path) -> TigerSummary:
    """解析老虎年度账单 PDF，提取已实现损益汇总

    PDF结构（按页码）：
    - Pages 5-47: 期权交易
    - Pages 48-50: 基金交易
    - Pages 50-59: 股票交易
    - Pages 59-65: 合约行权/过期
    - Page 68: 股息
    """
    summary = TigerSummary()

    with pdfplumber.open(filepath) as pdf:
        # 按section分段解析，使用页码定位
        all_lines = []
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_lines.extend(text.split("\n"))

        # 找到所有 "合计" 行（非个股合计，即section汇总行）
        # Section汇总行的格式: "合计 [数字开头]" 而非 "合计 [字母/代码]"
        parse_tiger_by_sections(all_lines, summary)
        parse_tiger_dividends_from_lines(all_lines, summary)

    return summary


def parse_tiger_by_sections(lines: list[str], summary: TigerSummary):
    """按section解析老虎PDF的合计行

    PDF中section标题为独立行: "期权", "外汇", "基金", "股票", "合约行权/过期"
    Section汇总行格式: "合计 数字 数字 ... 数字 HKD/USD"
    个股合计行格式: "合计 代码 ..." (含股票/期权代码)
    """
    current_section = ""
    # 用于避免后续重复section标题干扰（期末持仓等也有"期权"/"股票"子标题）
    sections_found = set()

    for i, line in enumerate(lines):
        stripped = line.strip()

        # 识别独立的section标题行
        if stripped == "期权" and "期权" not in sections_found:
            current_section = "期权"
            sections_found.add("期权")
            continue
        if stripped == "外汇":
            current_section = "外汇"
            continue
        if stripped == "基金" and "基金" not in sections_found:
            current_section = "基金"
            sections_found.add("基金")
            continue
        if stripped == "股票" and "股票" not in sections_found:
            current_section = "股票"
            sections_found.add("股票")
            continue
        if stripped == "合约行权/过期":
            current_section = "行权"
            sections_found.add("行权")
            continue
        if "入金与出金" in stripped:
            current_section = ""
            continue
        if "期末持仓" in stripped:
            current_section = ""
            continue

        # 跳过基础币种行和非合计行
        if "合计（基础币种）" in line:
            continue
        if "合计" not in line:
            continue
        if not current_section or current_section == "外汇":
            continue

        # 判断是section汇总还是个股汇总
        after_total = line.split("合计", 1)[1].strip()
        if not after_total:
            continue

        # 个股汇总: "合计 CODE num num ..." (CODE是字母或纯整数如02459)
        # Section汇总: "合计 -342.00 ..." (第一个token含小数点或千位逗号)
        first_token = after_total.split()[0] if after_total.split() else ""
        # Section total: starts with a number containing decimal point or comma
        is_section_total = bool(re.match(r'^-?[\d,]*\.\d', first_token))

        if not is_section_total:
            continue

        # 这是section汇总行
        currency_match = re.search(r"(HKD|USD)\s*$", line)
        if not currency_match:
            continue
        currency = currency_match.group(1)

        numbers = re.findall(r"([-]?[\d,]+\.?\d*)", line)
        if not numbers:
            continue

        realized_pnl = parse_number(numbers[-1])

        # 根据section和币种分配（只取第一次出现的section汇总）
        if current_section == "期权":
            if currency == "HKD" and summary.option_trade_hk == 0:
                summary.option_trade_hk = realized_pnl
            elif currency == "USD" and summary.option_trade_us == 0:
                summary.option_trade_us = realized_pnl
        elif current_section == "股票":
            if currency == "HKD" and summary.stock_hk == 0:
                summary.stock_hk = realized_pnl
            elif currency == "USD" and summary.stock_us == 0:
                summary.stock_us = realized_pnl
        elif current_section == "基金":
            if currency == "HKD" and summary.fund_hk == 0:
                summary.fund_hk = realized_pnl
            elif currency == "USD" and summary.fund_us == 0:
                summary.fund_us = realized_pnl
        elif current_section == "行权":
            if currency == "HKD" and summary.option_exercise_hk == 0:
                summary.option_exercise_hk = realized_pnl
            elif currency == "USD" and summary.option_exercise_us == 0:
                summary.option_exercise_us = realized_pnl


def parse_tiger_dividends_from_lines(lines: list[str], summary: TigerSummary):
    """解析分红记录 — 通用多行解析，支持多只股票

    老虎PDF分红格式为多行：
      Line 1: "中文名 数量：N"
      Line 2: "日期 股票 派发 gross 0 分红税: tax net currency"
      Line 3: "(CODE) 股息率：..."
    """
    in_dividend_section = False
    current_name = ""
    current_date = None

    for i, line in enumerate(lines):
        if "股息" in line and not in_dividend_section:
            in_dividend_section = True
            continue
        if in_dividend_section and ("补贴" in line or "奖励" in line or "在途" in line):
            in_dividend_section = False
            continue
        if in_dividend_section and "合计（基础币种）" in line:
            in_dividend_section = False
            continue

        if not in_dividend_section:
            continue

        # 匹配名称行: "英伟达 数量：100"
        name_match = re.match(r'^([\u4e00-\u9fff]+)\s+数量[：:](\d+)', line)
        if name_match:
            current_name = name_match.group(1)
            continue

        # 匹配数据行: "2025-10-02 股票 派发 1.00 0 分红税: 0.10 0.90 USD"
        data_match = re.match(r'^(\d{4}-\d{2}-\d{2})\s+.*?派发\s+', line)
        if data_match:
            date_str = data_match.group(1)
            try:
                current_date = datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                current_date = datetime(2025, 1, 1)

            # 提取数字：gross, tax, net
            numbers = re.findall(r"([-]?[\d,]+\.\d+)", line)
            # 检测币种
            currency = "USD"
            if "HKD" in line:
                currency = "HKD"

            # 提取分红税
            tax = 0.0
            tax_match = re.search(r'分红税:\s*([\d.]+)', line)
            if tax_match:
                tax = float(tax_match.group(1))

            # gross 是 "派发" 后第一个数字
            gross = 0.0
            if numbers:
                gross = parse_number(numbers[0])

            continue

        # 匹配代码行: "(NVDA) 股息率：..." 或 "(00700) 股息率：..."
        code_match = re.match(r'^\(([A-Z0-9]+)\)\s*股息率', line)
        if code_match and current_name:
            ticker = code_match.group(1)
            # 回溯找到上面解析的 gross/tax/currency/date
            # 需要从前面行取得数据，用局部变量
            # 重新扫描前一行获取数据
            prev_line = lines[i - 1] if i > 0 else ""
            prev_data_match = re.match(r'^(\d{4}-\d{2}-\d{2})\s+.*?派发\s+', prev_line)
            if prev_data_match:
                date_str = prev_data_match.group(1)
                try:
                    div_date = datetime.strptime(date_str, "%Y-%m-%d")
                except ValueError:
                    div_date = datetime(2025, 1, 1)

                numbers = re.findall(r"([-]?[\d,]+\.\d+)", prev_line)
                currency = "USD" if "USD" in prev_line else "HKD"

                tax = 0.0
                tax_match = re.search(r'分红税:\s*([\d.]+)', prev_line)
                if tax_match:
                    tax = float(tax_match.group(1))

                gross = parse_number(numbers[0]) if numbers else 0.0

                summary.dividends.append(Dividend(
                    code=ticker,
                    name=current_name,
                    amount=gross,
                    currency=currency,
                    date=div_date,
                    tax_withheld=tax,
                ))
            current_name = ""
            continue

    # Fallback: 如果未解析到任何分红记录
    if not summary.dividends:
        summary.dividends.append(Dividend(
            code="UNKNOWN",
            name="未识别",
            amount=0.0,
            currency="USD",
            date=datetime(2025, 1, 1),
            tax_withheld=0.0,
        ))
        print("  [警告] 未能从PDF解析到分红记录，请检查PDF结构是否已变化")


def parse_number(s: str) -> float:
    """解析数字字符串，处理千位逗号和负号"""
    s = s.replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# ============================================================
# 税额计算
# ============================================================


def to_cny(amount: float, currency: str, exchange_rates: dict) -> float:
    """按年度平均汇率折算为人民币"""
    rate = exchange_rates.get(currency.upper(), 1.0)
    return amount * rate


def calculate_tax(
    futu_gains: list[RealizedGain],
    futu_dividends: list[Dividend],
    tiger: TigerSummary,
    exchange_rates: dict,
) -> dict:
    """计算应纳税额"""

    result = {
        "futu": {},
        "tiger": {},
        "summary": {},
    }

    # ============ 富途 ============
    # 按类别和市场汇总
    futu_stock_gains_hkd = sum(
        g.gain for g in futu_gains if g.category == "股票"
    )
    futu_option_gains_hkd = sum(
        g.gain for g in futu_gains if g.category == "期权"
    )

    futu_stock_gains_cny = to_cny(futu_stock_gains_hkd, "HKD", exchange_rates)
    futu_option_gains_cny = to_cny(futu_option_gains_hkd, "HKD", exchange_rates)
    futu_total_gains_cny = futu_stock_gains_cny + futu_option_gains_cny

    futu_dividend_total_hkd = sum(d.amount for d in futu_dividends)
    futu_dividend_cny = to_cny(futu_dividend_total_hkd, "HKD", exchange_rates)
    # H股分红的10%预扣税（属于中国国内税，非境外税）
    futu_hshare_tax_hkd = sum(d.tax_withheld for d in futu_dividends)
    futu_hshare_tax_cny = to_cny(futu_hshare_tax_hkd, "HKD", exchange_rates)

    result["futu"] = {
        "stock_gains_hkd": futu_stock_gains_hkd,
        "option_gains_hkd": futu_option_gains_hkd,
        "total_gains_hkd": futu_stock_gains_hkd + futu_option_gains_hkd,
        "stock_gains_cny": futu_stock_gains_cny,
        "option_gains_cny": futu_option_gains_cny,
        "total_gains_cny": futu_total_gains_cny,
        "dividend_hkd": futu_dividend_total_hkd,
        "dividend_cny": futu_dividend_cny,
        "hshare_tax_hkd": futu_hshare_tax_hkd,
        "hshare_tax_cny": futu_hshare_tax_cny,
    }

    # ============ 老虎 ============
    # 港股部分（HKD）
    tiger_hk_option = tiger.option_trade_hk + tiger.option_exercise_hk
    tiger_hk_stock = tiger.stock_hk
    tiger_hk_total = tiger_hk_option + tiger_hk_stock
    tiger_hk_total_cny = to_cny(tiger_hk_total, "HKD", exchange_rates)

    # 美股部分（USD）
    tiger_us_option = tiger.option_trade_us + tiger.option_exercise_us
    tiger_us_stock = tiger.stock_us
    tiger_us_total = tiger_us_option + tiger_us_stock
    tiger_us_total_cny = to_cny(tiger_us_total, "USD", exchange_rates)

    # 老虎分红
    tiger_dividend_usd = sum(d.amount for d in tiger.dividends)
    tiger_dividend_cny = to_cny(tiger_dividend_usd, "USD", exchange_rates)
    tiger_tax_withheld_usd = sum(d.tax_withheld for d in tiger.dividends)
    tiger_tax_withheld_cny = to_cny(tiger_tax_withheld_usd, "USD", exchange_rates)

    result["tiger"] = {
        "hk_option_hkd": tiger_hk_option,
        "hk_stock_hkd": tiger_hk_stock,
        "hk_total_hkd": tiger_hk_total,
        "hk_total_cny": tiger_hk_total_cny,
        "us_option_usd": tiger_us_option,
        "us_stock_usd": tiger_us_stock,
        "us_total_usd": tiger_us_total,
        "us_total_cny": tiger_us_total_cny,
        "dividend_usd": tiger_dividend_usd,
        "dividend_cny": tiger_dividend_cny,
        "tax_withheld_usd": tiger_tax_withheld_usd,
        "tax_withheld_cny": tiger_tax_withheld_cny,
    }

    # ============ 汇总计算 ============
    # 按地区分别计算（中国税法要求分国别计算抵免限额）

    # 香港来源所得（富途港股 + 老虎港股）
    hk_capital_gains_cny = futu_total_gains_cny + tiger_hk_total_cny
    hk_dividend_cny = futu_dividend_cny  # 富途分红均为港股

    # 美国来源所得（老虎美股）
    us_capital_gains_cny = tiger_us_total_cny
    us_dividend_cny = tiger_dividend_cny

    # 资本利得税（盈亏可相抵，但不同国家分开）
    # 注：根据税法实务，同一国家/地区的资本利得可以盈亏相抵
    hk_gains_taxable = max(0, hk_capital_gains_cny)
    us_gains_taxable = max(0, us_capital_gains_cny)

    # 也计算合并后的情况（部分地区税务实务中可合并计算）
    total_capital_gains_cny = hk_capital_gains_cny + us_capital_gains_cny
    total_gains_taxable = max(0, total_capital_gains_cny)

    # 股息税
    total_dividend_cny = hk_dividend_cny + us_dividend_cny

    # 应纳税额
    tax_capital_gains_separate = (hk_gains_taxable + us_gains_taxable) * TAX_RATE
    tax_capital_gains_combined = total_gains_taxable * TAX_RATE
    tax_dividend = total_dividend_cny * TAX_RATE
    total_tax_separate = tax_capital_gains_separate + tax_dividend
    total_tax_combined = tax_capital_gains_combined + tax_dividend

    # 境外已缴税款抵扣（美股预扣税）
    foreign_tax_credit_cny = tiger_tax_withheld_cny
    # 抵扣限额：美国来源所得对应的应纳税额
    us_tax_limit = (max(0, us_capital_gains_cny) + us_dividend_cny) * TAX_RATE
    actual_credit = min(foreign_tax_credit_cny, us_tax_limit)

    # H股分红已缴的中国国内10%税（可抵减应纳税额）
    hshare_tax_credit = futu_hshare_tax_cny

    # 实际应缴（含H股已缴税款抵减）
    actual_tax_separate = max(0, total_tax_separate - actual_credit - hshare_tax_credit)
    actual_tax_combined = max(0, total_tax_combined - actual_credit - hshare_tax_credit)

    result["summary"] = {
        "hk_capital_gains_cny": hk_capital_gains_cny,
        "us_capital_gains_cny": us_capital_gains_cny,
        "total_capital_gains_cny": total_capital_gains_cny,
        "hk_gains_taxable": hk_gains_taxable,
        "us_gains_taxable": us_gains_taxable,
        "total_gains_taxable": total_gains_taxable,
        "hk_dividend_cny": hk_dividend_cny,
        "us_dividend_cny": us_dividend_cny,
        "total_dividend_cny": total_dividend_cny,
        "tax_capital_gains_separate": tax_capital_gains_separate,
        "tax_capital_gains_combined": tax_capital_gains_combined,
        "tax_dividend": tax_dividend,
        "total_tax_separate": total_tax_separate,
        "total_tax_combined": total_tax_combined,
        "foreign_tax_credit_cny": foreign_tax_credit_cny,
        "us_tax_limit": us_tax_limit,
        "actual_credit": actual_credit,
        "hshare_tax_credit": hshare_tax_credit,
        "actual_tax_separate": actual_tax_separate,
        "actual_tax_combined": actual_tax_combined,
    }

    return result


# ============================================================
# 报告生成
# ============================================================


def extract_option_underlying(code: str) -> str:
    """从期权代码中提取标的资产代码
    例: TCH250320C480000 → TCH, CSP250627P7750 → CSP
    """
    match = re.match(r'^([A-Z]+)\d', code)
    if match:
        return match.group(1)
    return code[:6] if len(code) > 6 else code


def generate_report(
    futu_gains: list[RealizedGain],
    futu_dividends: list[Dividend],
    tiger: TigerSummary,
    tax_result: dict,
    exchange_rates: dict,
    year: int = 2025,
) -> str:
    """生成税务计算报告"""

    lines = []
    lines.append("=" * 70)
    lines.append(f"       {year}年度境外证券投资所得个人所得税计算报告")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"汇率 (年度平均): USD/CNY = {exchange_rates['USD']}, "
                 f"HKD/CNY = {exchange_rates['HKD']}")
    lines.append("")

    # ============ 一、富途账户 ============
    lines.append("-" * 70)
    lines.append("一、富途证券账户 (港股)")
    lines.append("-" * 70)
    lines.append("")

    # 股票损益明细
    stock_gains = [g for g in futu_gains if g.category == "股票"]
    option_gains = [g for g in futu_gains if g.category == "期权"]

    lines.append("1. 股票已实现损益:")
    lines.append(f"   {'代码':<10} {'名称':<20} {'损益(HKD)':>12}")
    lines.append(f"   {'-'*10} {'-'*20} {'-'*12}")

    # 按代码汇总
    stock_by_code = {}
    for g in stock_gains:
        key = g.code
        if key not in stock_by_code:
            stock_by_code[key] = {"name": g.name, "gain": 0}
        stock_by_code[key]["gain"] += g.gain

    for code in sorted(stock_by_code.keys()):
        info = stock_by_code[code]
        lines.append(f"   {code:<10} {info['name']:<20} {info['gain']:>12.2f}")

    futu = tax_result["futu"]
    lines.append(f"   {'':10} {'合计':<20} {futu['stock_gains_hkd']:>12.2f}")
    lines.append(f"   折合人民币: {futu['stock_gains_cny']:.2f} CNY")
    lines.append("")

    # 期权损益
    lines.append("2. 期权已实现损益:")
    lines.append(f"   {'代码':<10} {'名称':<20} {'损益(HKD)':>12}")
    lines.append(f"   {'-'*10} {'-'*20} {'-'*12}")

    option_by_underlying = {}
    for g in option_gains:
        key = extract_option_underlying(g.code)
        if key not in option_by_underlying:
            option_by_underlying[key] = {"name": g.name, "gain": 0}
        option_by_underlying[key]["gain"] += g.gain

    for code in sorted(option_by_underlying.keys()):
        info = option_by_underlying[code]
        display_name = info["name"][:18] if len(info["name"]) > 18 else info["name"]
        lines.append(f"   {code:<10} {display_name:<20} {info['gain']:>12.2f}")

    lines.append(f"   {'':10} {'合计':<20} {futu['option_gains_hkd']:>12.2f}")
    lines.append(f"   折合人民币: {futu['option_gains_cny']:.2f} CNY")
    lines.append("")

    # 分红
    lines.append("3. 股息/分红收入:")
    lines.append(f"   {'代码':<10} {'名称':<20} {'金额(HKD)':>12} {'日期'}")
    lines.append(f"   {'-'*10} {'-'*20} {'-'*12} {'-'*10}")
    for d in futu_dividends:
        lines.append(
            f"   {d.code:<10} {d.name:<20} {d.amount:>12.2f} "
            f"{d.date.strftime('%Y-%m-%d')}"
        )
    lines.append(f"   {'':10} {'合计':<20} {futu['dividend_hkd']:>12.2f}")
    lines.append(f"   折合人民币: {futu['dividend_cny']:.2f} CNY")
    lines.append("")

    # 富途小计
    lines.append(f"   富途账户合计资本利得: {futu['total_gains_hkd']:.2f} HKD "
                 f"= {futu['total_gains_cny']:.2f} CNY")
    lines.append("")

    # ============ 二、老虎账户 ============
    lines.append("-" * 70)
    lines.append("二、老虎证券账户 (港股 + 美股)")
    lines.append("-" * 70)
    lines.append("")

    tiger_r = tax_result["tiger"]

    lines.append("1. 港股部分 (HKD):")
    lines.append(f"   期权交易已实现损益:     {tiger.option_trade_hk:>12.2f} HKD")
    lines.append(f"   期权行权/过期损益:      {tiger.option_exercise_hk:>12.2f} HKD")
    lines.append(f"   股票已实现损益:         {tiger.stock_hk:>12.2f} HKD")
    lines.append(f"   港股合计:               {tiger_r['hk_total_hkd']:>12.2f} HKD")
    lines.append(f"   折合人民币:             {tiger_r['hk_total_cny']:>12.2f} CNY")
    lines.append("")

    lines.append("2. 美股部分 (USD):")
    lines.append(f"   期权交易已实现损益:     {tiger.option_trade_us:>12.2f} USD")
    lines.append(f"   期权行权/过期损益:      {tiger.option_exercise_us:>12.2f} USD")
    lines.append(f"   股票已实现损益:         {tiger.stock_us:>12.2f} USD")
    lines.append(f"   美股合计:               {tiger_r['us_total_usd']:>12.2f} USD")
    lines.append(f"   折合人民币:             {tiger_r['us_total_cny']:>12.2f} CNY")
    lines.append("")

    lines.append("3. 分红收入:")
    for d in tiger.dividends:
        lines.append(
            f"   {d.code} {d.name}: 分红 {d.amount:.2f} {d.currency}, "
            f"预扣税 {d.tax_withheld:.2f} {d.currency}"
        )
    lines.append(f"   分红折合人民币: {tiger_r['dividend_cny']:.2f} CNY")
    lines.append(f"   已缴境外税款折合: {tiger_r['tax_withheld_cny']:.2f} CNY")
    lines.append("")

    lines.append("4. 基金收益 (货币基金，仅供参考):")
    lines.append(f"   港元货币基金: {tiger.fund_hk:.2f} HKD")
    lines.append(f"   美元流动基金: {tiger.fund_us:.2f} USD")
    lines.append("   注: 货币基金收益是否需申报存在争议，本报告暂不计入应税所得")
    lines.append("")

    # ============ 三、汇总 ============
    lines.append("-" * 70)
    lines.append("三、应纳税额计算")
    lines.append("-" * 70)
    lines.append("")

    s = tax_result["summary"]

    lines.append("1. 资本利得汇总 (折合人民币):")
    lines.append(f"   香港来源: {s['hk_capital_gains_cny']:>12.2f} CNY "
                 f"(富途港股 + 老虎港股)")
    lines.append(f"   美国来源: {s['us_capital_gains_cny']:>12.2f} CNY "
                 f"(老虎美股)")
    lines.append(f"   合计:     {s['total_capital_gains_cny']:>12.2f} CNY")
    lines.append("")

    lines.append("2. 股息收入汇总 (折合人民币):")
    lines.append(f"   香港来源: {s['hk_dividend_cny']:>12.2f} CNY")
    lines.append(f"   美国来源: {s['us_dividend_cny']:>12.2f} CNY")
    lines.append(f"   合计:     {s['total_dividend_cny']:>12.2f} CNY")
    lines.append("")

    lines.append("3. 应纳税额计算:")
    lines.append("")
    lines.append("   方案A: 分国家/地区分别计算 (严格按税法)")
    lines.append(f"   香港资本利得应税所得: max(0, {s['hk_capital_gains_cny']:.2f}) "
                 f"= {s['hk_gains_taxable']:.2f} CNY")
    lines.append(f"   美国资本利得应税所得: max(0, {s['us_capital_gains_cny']:.2f}) "
                 f"= {s['us_gains_taxable']:.2f} CNY")
    lines.append(f"   资本利得税 = ({s['hk_gains_taxable']:.2f} + "
                 f"{s['us_gains_taxable']:.2f}) × 20% "
                 f"= {s['tax_capital_gains_separate']:.2f} CNY")
    lines.append(f"   股息税 = {s['total_dividend_cny']:.2f} × 20% "
                 f"= {s['tax_dividend']:.2f} CNY")
    lines.append(f"   应纳税额合计 = {s['total_tax_separate']:.2f} CNY")
    lines.append("")
    lines.append("   方案B: 合并计算 (盈亏相抵)")
    lines.append(f"   资本利得应税所得: max(0, {s['total_capital_gains_cny']:.2f}) "
                 f"= {s['total_gains_taxable']:.2f} CNY")
    lines.append(f"   资本利得税 = {s['total_gains_taxable']:.2f} × 20% "
                 f"= {s['tax_capital_gains_combined']:.2f} CNY")
    lines.append(f"   股息税 = {s['total_dividend_cny']:.2f} × 20% "
                 f"= {s['tax_dividend']:.2f} CNY")
    lines.append(f"   应纳税额合计 = {s['total_tax_combined']:.2f} CNY")
    lines.append("")

    lines.append("4. 已缴税款抵扣:")
    lines.append(f"   美股股息预扣税 (境外): {s['foreign_tax_credit_cny']:.2f} CNY")
    lines.append(f"     抵扣限额 (美国来源所得应纳税): {s['us_tax_limit']:.2f} CNY")
    lines.append(f"     实际可抵扣: {s['actual_credit']:.2f} CNY")
    lines.append(f"   H股分红已缴税 (国内): {s['hshare_tax_credit']:.2f} CNY")
    lines.append(f"     (H股分红10%税由企业代扣代缴，属于已缴国内个税)")
    lines.append("")

    lines.append("-" * 70)
    lines.append("五、个税App申报填写参考")
    lines.append("-" * 70)
    lines.append("")

    # 计算分栏数据
    dividend_tax_gross = s['total_dividend_cny'] * TAX_RATE
    dividend_credit = s['actual_credit'] + s['hshare_tax_credit']
    dividend_tax_net = max(0, dividend_tax_gross - dividend_credit)
    capital_tax_a = s['tax_capital_gains_separate']
    capital_tax_b = s['tax_capital_gains_combined']

    lines.append("┌─────────────────────────────────────────────────────────────────┐")
    lines.append("│ 栏目：财产转让所得                                              │")
    lines.append("├─────────────────────────────────────────────────────────────────┤")
    lines.append(f"│  应纳税所得额 (方案A): {s['hk_gains_taxable'] + s['us_gains_taxable']:>12.2f} CNY              │")
    lines.append(f"│  应纳税所得额 (方案B): {s['total_gains_taxable']:>12.2f} CNY              │")
    lines.append(f"│  应纳税额     (方案A): {capital_tax_a:>12.2f} CNY              │")
    lines.append(f"│  应纳税额     (方案B): {capital_tax_b:>12.2f} CNY              │")
    lines.append(f"│  境外已缴税额:                0.00 CNY (资本利得无预扣税)  │")
    lines.append("└─────────────────────────────────────────────────────────────────┘")
    lines.append("")
    lines.append("┌─────────────────────────────────────────────────────────────────┐")
    lines.append("│ 栏目：利息、股息、红利所得                                      │")
    lines.append("├─────────────────────────────────────────────────────────────────┤")
    lines.append(f"│  应纳税所得额:         {s['total_dividend_cny']:>12.2f} CNY              │")
    lines.append(f"│  应纳税额:             {dividend_tax_gross:>12.2f} CNY              │")
    lines.append(f"│  境外已缴税额:         {s['foreign_tax_credit_cny']:>12.2f} CNY (美股预扣税)     │")
    lines.append(f"│  H股已缴税(国内代扣):  {s['hshare_tax_credit']:>12.2f} CNY              │")
    lines.append(f"│  实际应补缴:           {dividend_tax_net:>12.2f} CNY              │")
    lines.append("└─────────────────────────────────────────────────────────────────┘")
    lines.append("")
    lines.append("┌─────────────────────────────────────────────────────────────────┐")
    lines.append("│ 合计实际应缴                                                    │")
    lines.append("├─────────────────────────────────────────────────────────────────┤")
    lines.append(f"│  方案A (分国别): {s['actual_tax_separate']:>12.2f} CNY                       │")
    lines.append(f"│  方案B (合并):   {s['actual_tax_combined']:>12.2f} CNY                       │")
    lines.append("└─────────────────────────────────────────────────────────────────┘")
    lines.append("")

    # ============ 备注 ============
    lines.append("-" * 70)
    lines.append("六、重要说明")
    lines.append("-" * 70)
    lines.append("")
    lines.append("1. 本报告仅供参考，不构成税务建议。实际申报请咨询专业税务顾问。")
    lines.append("2. 汇率采用年度平均汇率，实际申报可能需要按交易日汇率计算。")
    lines.append("3. 富途账户使用加权平均成本法计算，期初持仓成本需用户确认。")
    lines.append("4. 老虎账户直接使用券商提供的已实现损益数据。")
    lines.append("5. 货币基金收益未计入应税所得，如需申报请另行计算。")
    lines.append("6. 中国税法对不同国家/地区的税收抵免有分国不分项的限额规定。")
    lines.append("")
    lines.append("\U0001f4cc 特别说明（为什么存在两个方案）：")
    lines.append('   根据中国个税法\u201c分国不分项\u201d的字面原则，不同国家的盈亏不应合并。')
    lines.append("   但在个税App实务申报中，部分地方税务局（如北京、上海等部分区域）出于")
    lines.append('   合理税负原则，允许纳税人在\u201c境外所得\u201d综合模块中合并冲抵同年盈亏。')
    lines.append("   最终应纳税额取决于您户籍地或纳税所在地主管税务局的实操口径，建议优先")
    lines.append("   尝试方案B（合并对冲），若系统或人工审核不通过，再按方案A（分国别）补缴。")
    lines.append("")
    lines.append("=" * 70)

    return "\n".join(lines)


# ============================================================
# 主程序
# ============================================================


def parse_args() -> argparse.Namespace:
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description="境外资本利得税计算工具 — 计算中国大陆税务居民境外证券投资所得的个人所得税",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
示例:
  python tax_calculator.py                         # 默认: 2025年度, 当前目录
  python tax_calculator.py --year 2026             # 2026年度
  python tax_calculator.py --futu my_futu.xlsx --tiger my_tiger.pdf
  python tax_calculator.py --usd-rate 7.25 --hkd-rate 0.93
""",
    )
    parser.add_argument(
        "year", type=int,
        help="税务年度 (如 2025)，用于推断文件名和报告标题",
    )
    parser.add_argument(
        "--futu", type=str, default=None,
        help="富途年度账单 Excel 文件路径 (默认: <year>_年度账单_futu.xlsx)",
    )
    parser.add_argument(
        "--tiger", type=str, default=None,
        help="老虎年度账单 PDF 文件路径 (默认: <year>-年度账单-tiger.pdf)",
    )
    parser.add_argument(
        "--usd-rate", type=float, default=None,
        help="USD/CNY 汇率 (覆盖默认值，也可用环境变量 EXCHANGE_RATE_USD)",
    )
    parser.add_argument(
        "--hkd-rate", type=float, default=None,
        help="HKD/CNY 汇率 (覆盖默认值，也可用环境变量 EXCHANGE_RATE_HKD)",
    )
    parser.add_argument(
        "--output-dir", type=str, default=None,
        help="输出目录 (默认: ./output)",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    year = args.year
    base_dir = Path(__file__).parent

    # 汇率配置（优先级: CLI参数 > 环境变量 > 默认值）
    exchange_rates = {
        "USD": args.usd_rate or float(os.environ.get("EXCHANGE_RATE_USD", "7.1886")),
        "HKD": args.hkd_rate or float(os.environ.get("EXCHANGE_RATE_HKD", "0.9243")),
        "CNY": 1.0,
        "RMB": 1.0,
    }

    # 文件路径
    futu_file = Path(args.futu) if args.futu else base_dir / f"{year}-年度账单-futu.xlsx"
    tiger_file = Path(args.tiger) if args.tiger else base_dir / f"{year}-年度账单-tiger.pdf"
    output_dir = Path(args.output_dir) if args.output_dir else base_dir / "output"
    output_file = output_dir / f"tax_report_{year}.txt"

    # 文件存在性检查
    if not futu_file.exists():
        print(f"[错误] 富途账单文件不存在: {futu_file}")
        print(f"  请确认文件路径正确，或使用 --futu 参数指定文件位置")
        sys.exit(1)
    if not tiger_file.exists():
        print(f"[错误] 老虎账单文件不存在: {tiger_file}")
        print(f"  请确认文件路径正确，或使用 --tiger 参数指定文件位置")
        sys.exit(1)

    print(f"境外资本利得税计算工具 v2.0")
    print(f"税务年度: {year}")
    print(f"数据源: {futu_file.name}, {tiger_file.name}")
    print()

    # 1. 解析富途数据
    print("[1/4] 解析富途年度账单...")
    futu_trades, futu_dividends, initial_positions = parse_futu_trades(futu_file)
    print(f"      交易记录: {len(futu_trades)} 笔")
    print(f"      分红记录: {len(futu_dividends)} 笔")
    if initial_positions:
        print(f"      期初持仓: {len(initial_positions)} 只")
        for code, pos in initial_positions.items():
            print(f"        {code}: {pos.quantity:.0f}股, "
                  f"成本 {pos.total_cost:.2f} HKD "
                  f"(均价 {pos.avg_cost:.4f})")

    # 2. 计算富途损益
    print("[2/4] 计算富途已实现损益 (加权平均成本法)...")
    futu_gains, futu_positions = calculate_futu_gains(futu_trades, initial_positions)
    print(f"      已实现损益笔数: {len(futu_gains)} 笔")

    stock_total = sum(g.gain for g in futu_gains if g.category == "股票")
    option_total = sum(g.gain for g in futu_gains if g.category == "期权")
    print(f"      股票损益: {stock_total:.2f} HKD")
    print(f"      期权损益: {option_total:.2f} HKD")

    # 3. 解析老虎数据
    print("[3/4] 解析老虎年度账单...")
    tiger = parse_tiger_pdf(tiger_file)
    print(f"      期权(交易): HK {tiger.option_trade_hk:.2f} HKD, "
          f"US {tiger.option_trade_us:.2f} USD")
    print(f"      期权(行权): HK {tiger.option_exercise_hk:.2f} HKD, "
          f"US {tiger.option_exercise_us:.2f} USD")
    print(f"      股票: HK {tiger.stock_hk:.2f} HKD, "
          f"US {tiger.stock_us:.2f} USD")
    print(f"      分红: {len(tiger.dividends)} 笔")

    # 检查解析结果是否全零（可能PDF结构已变）
    all_zero = (
        tiger.option_trade_hk == 0 and tiger.option_trade_us == 0 and
        tiger.option_exercise_hk == 0 and tiger.option_exercise_us == 0 and
        tiger.stock_hk == 0 and tiger.stock_us == 0
    )
    if all_zero:
        print("  [警告] 老虎PDF解析结果全部为零！可能PDF结构已变化，请人工核实。")

    # 4. 计算税额
    print("[4/4] 计算应纳税额...")
    tax_result = calculate_tax(futu_gains, futu_dividends, tiger, exchange_rates)

    # 5. 生成报告
    report = generate_report(futu_gains, futu_dividends, tiger, tax_result, exchange_rates, year)

    # 输出到终端
    print()
    print(report)

    # 保存到文件
    output_dir.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"\n报告已保存至: {output_file}")


if __name__ == "__main__":
    main()
